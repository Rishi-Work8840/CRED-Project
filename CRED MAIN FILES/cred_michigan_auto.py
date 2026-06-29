"""
CRED Accelerator — Michigan Auto Real-Data Pipeline
======================================================
Ingests Power BI metadata exports (17 CSV files in /Raw Metadata/),
joins them into the unified CRED schema, runs the scoring engine, and
produces an executive-ready Excel report.

Input:  Raw Metadata/*.csv  (Power BI Admin/Scanner API exports)
Output: outputs/CRED_MichiganAuto_Output.xlsx

Author: Migration Accelerator Team
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# Reuse the proven scoring engine
from cred_scoring_engine import (
    CONFIG,
    build_fingerprints,
    score_recency_usage,
    compute_pairwise_similarity,
    apply_cred_classification,
    prepare_outputs,
    run_sensitivity_analysis,
    load_and_validate,
)


# ─────────────────────────────────────────────────────────────
# SECTION 1 — CONFIGURATION
# ─────────────────────────────────────────────────────────────

# Code lives in CRED MAIN FILES/, but data lives at project root (one level up)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR      = os.path.join(PROJECT_ROOT, "Raw Metadata")
OUTPUT_DIR   = os.path.join(PROJECT_ROOT, "outputs")
OUTPUT_FILE  = os.path.join(OUTPUT_DIR, "CRED_MichiganAuto_Output.xlsx")

# Treat today as the snapshot reference date (recency is computed from this)
SNAPSHOT_DATE = pd.Timestamp.today().normalize()

# Workspace name → business domain mapping
WORKSPACE_DOMAIN_MAP = {
    "Michigan Auto - Production":     "Production",
    "Michigan Auto - FAR Reporting":  "Finance",
    "Michigan Auto - 8:15 KPIs":      "Operations",
    "Michigan Auto - Maintenance":    "Maintenance",
    "Michigan Auto - Data Science":   "Data Science",
    "Michigan Auto - Finance":        "Finance",
    "Michigan Auto - Executive":      "Executive",
    "Michigan Auto - Dev":            "Dev/Test",
    "Michigan Auto - UAT":            "Dev/Test",
    "Michigan Auto - Supply Chain":   "Supply Chain",
    "Michigan Auto - HR Analytics":   "HR",
    "Michigan Auto - Archive":        "Archive",
}


# ─────────────────────────────────────────────────────────────
# SECTION 2 — RAW FILE LOADER
# ─────────────────────────────────────────────────────────────

def load_raw_files() -> dict:
    """Load all CSVs from the Raw Metadata folder into a dict of DataFrames."""
    print("[INGEST] Loading raw metadata files...")
    files = {
        "workspaces":   "01_workspaces.csv",
        "datasets":     "02_datasets.csv",
        "reports":      "03_reports.csv",
        "report_pages": "04_report_pages.csv",
        "measures":     "05_measures.csv",
        "users":        "06_users.csv",
        "permissions":  "07_user_workspace_permissions.csv",
        "refresh":      "08_refresh_history.csv",
        "activity":     "09_activity_log.csv",
        "lineage":      "10_data_lineage.csv",
    }
    raw = {}
    for key, fname in files.items():
        path = os.path.join(RAW_DIR, fname)
        raw[key] = pd.read_csv(path)
        print(f"  + {fname:35s} {len(raw[key]):>6} rows")
    return raw


# ─────────────────────────────────────────────────────────────
# SECTION 3 — TRANSFORM TO CRED SCHEMA
# ─────────────────────────────────────────────────────────────

def derive_business_domain(workspace_name: str) -> str:
    """Map a workspace name to a business domain."""
    return WORKSPACE_DOMAIN_MAP.get(workspace_name, "Other")


def derive_owner_email(report_row, permissions_df) -> str:
    """
    Pick the most likely owner: the report's last_modified_by if present,
    otherwise the first Admin user on the report's workspace.
    """
    last_mod = report_row.get("last_modified_by")
    if isinstance(last_mod, str) and last_mod.strip():
        return last_mod.strip().lower()

    ws_admins = permissions_df[
        (permissions_df["workspace_id"] == report_row["workspace_id"]) &
        (permissions_df["permission_level"] == "Admin") &
        (permissions_df["is_active"] == True)
    ]
    if len(ws_admins) > 0:
        return ws_admins.iloc[0]["user_email"].lower()
    return "unknown@michiganauto.com"


def derive_report_type(report_row) -> str:
    """Map endorsement + workspace to a report_type."""
    ws_name = str(report_row.get("workspace_name", "")).lower()
    endorse = str(report_row.get("endorsement", "")).strip()

    if "uat" in ws_name or "dev" in ws_name:
        return "ad-hoc"
    if endorse == "Certified":
        return "executive"
    if endorse == "Promoted":
        return "scheduled"
    return "operational"


def build_metrics_per_dataset(measures_df: pd.DataFrame) -> dict:
    """For each dataset_id, return a pipe-delimited string of its measure names."""
    grouped = (
        measures_df[measures_df["is_hidden"] == False]
        .groupby("dataset_id")["measure_name"]
        .apply(lambda s: "|".join(sorted(set(s.str.lower().str.strip()))))
    )
    return grouped.to_dict()


def build_tables_per_dataset(lineage_df: pd.DataFrame) -> dict:
    """For each dataset_id, return a pipe-delimited string of its source tables/views."""
    def extract_table(obj: str) -> str:
        # source_object looks like 'MA_FAR_Gold.dbo.vw_FAR_Semantic_Model'
        # Take the last segment (the actual table/view name) and normalise.
        if not isinstance(obj, str) or not obj:
            return ""
        last = obj.split(".")[-1].lower().strip()
        return last

    lineage_df = lineage_df.copy()
    lineage_df["_table"] = lineage_df["source_object"].apply(extract_table)
    grouped = (
        lineage_df.groupby("dataset_id")["_table"]
        .apply(lambda s: "|".join(sorted(set(t for t in s if t))))
    )
    return grouped.to_dict()


def build_dimensions_per_report(pages_df: pd.DataFrame) -> dict:
    """
    Use report_pages info as a proxy for dimensions. Each report's
    'dimensions' becomes the pipe-delimited list of page names that
    have slicers (slicer pages reveal what the report breaks data down by).
    """
    slicer_pages = pages_df[pages_df["has_slicers"] == True].copy()
    slicer_pages["page_name"] = slicer_pages["page_name"].str.lower().str.strip()
    grouped = (
        slicer_pages.groupby("report_id")["page_name"]
        .apply(lambda s: "|".join(sorted(set(s))))
    )
    return grouped.to_dict()


def build_usage_from_activity(activity_df: pd.DataFrame) -> pd.DataFrame:
    """
    Recompute view_count_90d and last_accessed_date from the activity log
    (more reliable than the metadata snapshot fields).
    """
    activity_df = activity_df.copy()
    activity_df["activity_datetime"] = pd.to_datetime(activity_df["activity_datetime"], errors="coerce")

    # Filter to ViewReport events in the last 90 days from snapshot
    cutoff = SNAPSHOT_DATE - pd.Timedelta(days=90)
    views = activity_df[
        (activity_df["activity_type"] == "ViewReport") &
        (activity_df["activity_datetime"] >= cutoff)
    ]

    view_count = views.groupby("report_id").size().rename("view_count_90d")
    last_access = (
        activity_df.groupby("report_id")["activity_datetime"]
        .max().rename("last_accessed_date")
    )
    return pd.concat([view_count, last_access], axis=1).reset_index()


def enrich_operational_signals(cred_df: pd.DataFrame, raw: dict) -> pd.DataFrame:
    """
    Compute 5 OPTIONAL operational signals from the raw Power BI metadata
    and join them onto the CRED input DataFrame. These power the enriched
    UX gap and the new Health gap in the engine.

    Signals (per report unless noted):
      refresh_fail_rate_30d   failed/total refreshes in last 30d, looked up by dataset_id
      activity_fail_rate_30d  is_success=False / total activity in last 30d
      edit_count_90d          count of EditReport events in last 90d
      mobile_view_share       share of ViewReport events from Mobile/Embedded clients (90d)
      is_mobile_ready         proxy from embed_url_present in 03_reports.csv

    Missing/empty inputs default to neutral values (0 / True for readiness) so
    the engine still produces clean output even when an aggregate is empty.
    """
    print("\n[ENRICH] Computing operational signals from raw metadata...")
    df = cred_df.copy()
    cutoff_30d = SNAPSHOT_DATE - pd.Timedelta(days=30)
    cutoff_90d = SNAPSHOT_DATE - pd.Timedelta(days=90)

    # ── (a) refresh_fail_rate_30d (per dataset, then mapped onto reports) ──
    refresh = raw["refresh"].copy()
    refresh["start_datetime"] = pd.to_datetime(refresh["start_datetime"], errors="coerce")
    last30 = refresh[refresh["start_datetime"] >= cutoff_30d]
    by_ds = last30.groupby("dataset_id").apply(
        lambda g: (g["status"] == "Failed").sum() / max(1, len(g))
    )
    ds_to_rate = by_ds.to_dict()
    rid_to_did = raw["reports"].set_index("report_id")["dataset_id"].to_dict()
    df["refresh_fail_rate_30d"] = df["report_id"].map(
        lambda rid: round(float(ds_to_rate.get(rid_to_did.get(rid), 0.0)), 3)
    )

    # ── (b) activity_fail_rate_30d (per report) ───────────────────────────
    activity = raw["activity"].copy()
    activity["activity_datetime"] = pd.to_datetime(activity["activity_datetime"], errors="coerce")
    act_last30 = activity[activity["activity_datetime"] >= cutoff_30d]
    rpt_fail = act_last30.groupby("report_id").apply(
        lambda g: (~g["is_success"].astype(bool)).sum() / max(1, len(g))
    )
    df["activity_fail_rate_30d"] = df["report_id"].map(rpt_fail.to_dict()).fillna(0.0).round(3)

    # ── (c) edit_count_90d ────────────────────────────────────────────────
    act_last90 = activity[activity["activity_datetime"] >= cutoff_90d]
    edits = act_last90[act_last90["activity_type"] == "EditReport"]
    edit_count = edits.groupby("report_id").size()
    df["edit_count_90d"] = df["report_id"].map(edit_count.to_dict()).fillna(0).astype(int)

    # ── (d) mobile_view_share (last 90d) ──────────────────────────────────
    views_90d = act_last90[act_last90["activity_type"] == "ViewReport"]
    by_rpt_client = views_90d.groupby(["report_id", "client_type"]).size().unstack(fill_value=0)
    if len(by_rpt_client) > 0:
        total = by_rpt_client.sum(axis=1)
        mobile = by_rpt_client.get("Mobile", 0) + by_rpt_client.get("Embedded", 0)
        share = (mobile / total).fillna(0.0)
        df["mobile_view_share"] = df["report_id"].map(share.to_dict()).fillna(0.0).round(3)
    else:
        df["mobile_view_share"] = 0.0

    # ── (e) is_mobile_ready (proxy = embed_url_present) ───────────────────
    rid_to_embed = raw["reports"].set_index("report_id")["embed_url_present"].to_dict()
    df["is_mobile_ready"] = df["report_id"].map(
        lambda rid: bool(rid_to_embed.get(rid, True))
    )

    print(f"  + enriched {len(df)} reports with 5 operational signals")
    return df


def transform_to_cred_input(raw: dict) -> pd.DataFrame:
    """Join all raw files into the flat DataFrame the CRED engine expects."""
    print("\n[TRANSFORM] Building unified CRED input DataFrame...")
    reports = raw["reports"].copy()

    # Domain from workspace
    reports["business_domain"] = reports["workspace_name"].apply(derive_business_domain)

    # Owner from permissions (admin) with fallback to last_modified_by
    reports["owner_email"] = reports.apply(
        lambda r: derive_owner_email(r, raw["permissions"]), axis=1
    )

    # Platform is fixed for this export
    reports["source_platform"] = "power bi"

    # Report type
    reports["report_type"] = reports.apply(derive_report_type, axis=1)

    # Tags = endorsement + workspace_name (light free-text bucket)
    reports["tags"] = (
        reports["endorsement"].fillna("none").astype(str).str.lower() + "|" +
        reports["workspace_name"].fillna("").str.lower()
    )

    # Description = report_name + dataset_name (best available text)
    reports["report_description"] = (
        reports["report_name"].astype(str) + " | dataset: " +
        reports["dataset_name"].fillna("").astype(str)
    )

    # ── Usage (recomputed from activity log) ───────────────
    usage = build_usage_from_activity(raw["activity"])
    reports = reports.merge(usage, on="report_id", how="left", suffixes=("_meta", ""))

    # Fall back to metadata view counts if activity log had nothing
    reports["view_count_90d"] = (
        reports["view_count_90d"].fillna(reports["views_last_30d"]).fillna(0).astype(int)
    )
    reports["last_accessed_date"] = pd.to_datetime(
        reports["last_accessed_date"].fillna(reports["last_modified_date"]),
        errors="coerce"
    )

    # ── Tables (from data_lineage joined via dataset_id) ───
    tables_map = build_tables_per_dataset(raw["lineage"])
    reports["source_tables"] = reports["dataset_id"].map(tables_map).fillna("")

    # ── Metrics (from measures joined via dataset_id) ──────
    metrics_map = build_metrics_per_dataset(raw["measures"])
    reports["metrics"] = reports["dataset_id"].map(metrics_map).fillna("")

    # ── Dimensions (from report_pages, slicer pages) ───────
    dims_map = build_dimensions_per_report(raw["report_pages"])
    reports["dimensions"] = reports["report_id"].map(dims_map).fillna("")

    # ── Final shape: keep only the CRED required columns ──
    cred_cols = [
        "report_id", "report_name", "report_description",
        "source_platform", "business_domain", "owner_email",
        "last_accessed_date", "view_count_90d",
        "source_tables", "metrics", "dimensions",
        "report_type", "tags",
    ]
    cred_df = reports[cred_cols].copy()

    # Normalise string columns (engine expects lowercase)
    str_cols = ["source_tables", "metrics", "dimensions", "tags",
                "report_name", "report_description", "business_domain",
                "source_platform", "report_type"]
    for col in str_cols:
        cred_df[col] = cred_df[col].fillna("").astype(str).str.lower().str.strip()

    # ── Enrich with 5 OPTIONAL operational signals ────────────────────────
    # These feed the enriched UX gap and the new Health gap. Safe defaults if
    # any source CSV is empty/missing for a given report.
    cred_df = enrich_operational_signals(cred_df, raw)

    print(f"  Built CRED input: {len(cred_df)} reports x {len(cred_df.columns)} cols")
    return cred_df, reports  # also return enriched reports for context sheets


# ─────────────────────────────────────────────────────────────
# SECTION 4 — RUN ENGINE (minimal wrapper, no global validator)
# ─────────────────────────────────────────────────────────────

def run_engine(cred_df: pd.DataFrame) -> tuple:
    """Run the scoring engine stages directly (bypasses global I/O messages)."""
    print("\n[ENGINE] Running CRED scoring pipeline...")
    df = cred_df.copy()

    # Light type coercion (mirrors load_and_validate, without raising on extras)
    df["last_accessed_date"] = pd.to_datetime(df["last_accessed_date"], errors="coerce")
    df["view_count_90d"] = pd.to_numeric(df["view_count_90d"], errors="coerce").fillna(0).astype(int)
    df["_has_access_date"] = df["last_accessed_date"].notna()

    df = build_fingerprints(df)
    print("  + fingerprints built")

    df = score_recency_usage(df)
    print("  + recency & usage scored")

    df = compute_pairwise_similarity(df)
    n_clusters = int(df["_cluster_id"].max()) + 1 if df["_cluster_id"].max() >= 0 else 0
    print(f"  + similarity computed - {n_clusters} consolidation clusters found")

    df = apply_cred_classification(df)
    print("  + CRED classification complete")

    # Sensitivity: per-report columns at multiple thresholds
    print("  + running sensitivity analysis (0.50 / 0.75 / 0.90)...")
    sensitivity_cols = run_sensitivity_analysis(cred_df)
    df = df.merge(sensitivity_cols, left_on="report_id", right_index=True, how="left")
    print("  + sensitivity columns merged")

    return prepare_outputs(df)


# ─────────────────────────────────────────────────────────────
# SECTION 5 — EXCEL WRITER (multi-sheet, mentor-ready)
# ─────────────────────────────────────────────────────────────

def write_excel_output(tech, exc, summ, raw, cred_input):
    """Write a clean multi-sheet Excel deliverable."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Build supplementary summary sheets ─────────────────────
    # By domain
    by_domain = (
        tech.groupby("business_domain")["cred_label"]
        .value_counts().unstack(fill_value=0).reset_index()
    )
    # By workspace (join back via cred_input → reports)
    ws_lookup = raw["reports"][["report_id", "workspace_name"]]
    tech_ws = tech.merge(ws_lookup, on="report_id", how="left")
    by_workspace = (
        tech_ws.groupby("workspace_name")["cred_label"]
        .value_counts().unstack(fill_value=0).reset_index()
    )
    # Consolidation cluster detail
    clustered = tech[tech["_cluster_id"] >= 0].copy()
    clustered = clustered.sort_values(["_cluster_id", "_is_cluster_anchor"], ascending=[True, False])

    # Build enriched cluster documentation columns
    cluster_enrichment = []
    for cid in sorted(clustered["_cluster_id"].unique()):
        members = clustered[clustered["_cluster_id"] == cid]
        anchor_row = members[members["_is_cluster_anchor"] == True]
        if anchor_row.empty:
            continue
        anchor_name = anchor_row.iloc[0]["report_name"]
        anchor_views = anchor_row.iloc[0]["view_count_90d"]
        anchor_tables_str = anchor_row.iloc[0].get("source_tables", "")
        anchor_metrics_str = anchor_row.iloc[0].get("metrics", "")
        anchor_dims_str = anchor_row.iloc[0].get("dimensions", "")
        anchor_tables = {t.strip() for t in str(anchor_tables_str).split("|") if t.strip()}
        anchor_metrics = {m.strip() for m in str(anchor_metrics_str).split("|") if m.strip()}
        anchor_dims = {d.strip() for d in str(anchor_dims_str).split("|") if d.strip()}

        all_member_names = members["report_name"].tolist()
        cluster_size = len(members)
        second_highest = members[members["_is_cluster_anchor"] != True]["view_count_90d"].max() if cluster_size > 1 else 0

        for idx, row in members.iterrows():
            is_anchor = row["_is_cluster_anchor"]
            member_tables = {t.strip() for t in str(row.get("source_tables", "")).split("|") if t.strip()}
            member_metrics = {m.strip() for m in str(row.get("metrics", "")).split("|") if m.strip()}
            member_dims = {d.strip() for d in str(row.get("dimensions", "")).split("|") if d.strip()}

            # Unique capabilities (what this member has that anchor doesn't)
            if is_anchor:
                unique_cap = "N/A (this IS the anchor)"
                anchor_reason = f"Highest usage in cluster ({anchor_views} views vs next-highest {int(second_highest)})"
            else:
                unique_t = member_tables - anchor_tables
                unique_m = member_metrics - anchor_metrics
                unique_d = member_dims - anchor_dims
                parts = []
                if unique_t:
                    parts.append(f"tables: {', '.join(sorted(unique_t))}")
                if unique_m:
                    parts.append(f"metrics: {', '.join(sorted(unique_m))}")
                if unique_d:
                    parts.append(f"dims: {', '.join(sorted(unique_d))}")
                unique_cap = "; ".join(parts) if parts else "None — safe to merge without data loss"
                anchor_reason = ""

            cluster_enrichment.append({
                "idx": idx,
                "anchor_report_name": anchor_name,
                "cluster_size": cluster_size,
                "all_cluster_members": " | ".join(all_member_names),
                "anchor_selection_reason": anchor_reason if is_anchor else f"Anchor '{anchor_name}' chosen for highest usage ({anchor_views} views)",
                "unique_capability_at_risk": unique_cap,
            })

    enrichment_df = pd.DataFrame(cluster_enrichment).set_index("idx")
    clustered = clustered.join(enrichment_df)

    clusters = (
        clustered[["_cluster_id", "_is_cluster_anchor", "report_id", "report_name",
          "business_domain", "owner_email", "view_count_90d", "_sim_score",
          "_sim_partner_name", "cred_label",
          "anchor_report_name", "cluster_size", "all_cluster_members",
          "anchor_selection_reason", "unique_capability_at_risk"]]
        .rename(columns={
            "_cluster_id": "cluster_id",
            "_is_cluster_anchor": "is_anchor",
            "_sim_score": "similarity_score",
            "_sim_partner_name": "most_similar_to",
        })
    )

    # Top-line KPIs
    kpi_rows = [
        ["Snapshot date",                SNAPSHOT_DATE.strftime("%Y-%m-%d")],
        ["Total reports analysed",       summ["total_reports"]],
        ["Consolidate (merge)",          summ["consolidate"]],
        ["Retain (migrate as-is)",       summ["retain"]],
        ["Eliminate (retire)",           summ["eliminate"]],
        ["Develop (redesign)",           summ["develop"]],
        ["Migration reduction (%)",      summ["estimated_migration_reduction_pct"]],
        ["Consolidation clusters found", summ["consolidation_clusters"]],
        ["Run timestamp",                summ["run_timestamp"]],
    ]
    kpi_df = pd.DataFrame(kpi_rows, columns=["Metric", "Value"])

    print(f"\n[EXPORT] Writing Excel: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        kpi_df.to_excel(writer,        index=False, sheet_name="1. Executive Summary")
        exc.to_excel(writer,           index=False, sheet_name="2. Executive View")
        tech.to_excel(writer,          index=False, sheet_name="3. Technical View")
        clusters.to_excel(writer,      index=False, sheet_name="4. Consolidation Clusters")
        by_domain.to_excel(writer,     index=False, sheet_name="5. Breakdown by Domain")
        by_workspace.to_excel(writer,  index=False, sheet_name="6. Breakdown by Workspace")
        cred_input.to_excel(writer,    index=False, sheet_name="7. CRED Input (Built)")

    # Light cosmetic formatting (column widths)
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)
        ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"  Excel saved with {len(wb.sheetnames)} sheets")


# ─────────────────────────────────────────────────────────────
# SECTION 6 — MAIN
# ─────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  CRED ACCELERATOR — Michigan Auto Real-Data Pipeline")
    print(f"  Snapshot date: {SNAPSHOT_DATE.strftime('%Y-%m-%d')}")
    print("=" * 70)

    raw = load_raw_files()
    cred_input, enriched_reports = transform_to_cred_input(raw)
    tech, exc, summ = run_engine(cred_input)

    # Console summary
    print("\n" + "=" * 70)
    print("  RESULTS SUMMARY")
    print("=" * 70)
    print(f"  Total reports:        {summ['total_reports']}")
    print(f"  Consolidate:          {summ['consolidate']}")
    print(f"  Retain:               {summ['retain']}")
    print(f"  Eliminate:            {summ['eliminate']}")
    print(f"  Develop:              {summ['develop']}")
    print(f"  Migration reduction:  ~{summ['estimated_migration_reduction_pct']}%")
    print(f"  Clusters identified:  {summ['consolidation_clusters']}")
    print("=" * 70)

    write_excel_output(tech, exc, summ, raw, cred_input)
    print(f"\n[DONE] Output ready: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
